"""Research cohort, phase outcomes, analysis and exports."""
import io

import pytest

from warfarin import research
from warfarin.db import db, read_db
from warfarin.time_utils import now


@pytest.fixture
def study_dates():
    """Two consecutive 30-day windows entirely in the past."""
    from datetime import timedelta

    from warfarin.time_utils import DATE_FMT, now_dt

    end = now_dt().date() - timedelta(days=1)
    endline_start = end - timedelta(days=29)
    baseline_end = endline_start - timedelta(days=1)
    baseline_start = baseline_end - timedelta(days=29)
    return {
        "baseline_start": baseline_start.strftime(DATE_FMT),
        "baseline_end": baseline_end.strftime(DATE_FMT),
        "endline_start": endline_start.strftime(DATE_FMT),
        "endline_end": end.strftime(DATE_FMT),
    }


def _enroll(patient, study_dates, arm="intervention", **overrides):
    form = {
        "arm": arm, "status": "enrolled", "consent_date": study_dates["baseline_start"],
        "consent_version": "v1.0", **study_dates,
    }
    form.update(overrides)
    with db() as conn:
        return research.enroll(conn, patient["patient_id"], form, "pytest")


def _seed_doses(patient_id, start, end, taken_fraction):
    """Fill a window with daily doses, marking a share of them taken."""
    from datetime import timedelta

    from warfarin.time_utils import DATE_FMT, parse_date

    current = parse_date(start)
    last = parse_date(end)
    index = 0
    with db() as conn:
        while current <= last:
            taken = (index % 10) < round(taken_fraction * 10)
            conn.execute(
                "INSERT INTO medication_plan (patient_id, scheduled_date, scheduled_time, "
                "warfarin_mg, status, created_at) VALUES (?,?,?,?,?,?)",
                (patient_id, current.strftime(DATE_FMT), "18:00", 3.0,
                 "taken" if taken else "missed", now()),
            )
            current += timedelta(days=1)
            index += 1


def _seed_labs(patient_id, dates_and_values):
    with db() as conn:
        for date_string, value in dates_and_values:
            conn.execute(
                "INSERT INTO lab_results (patient_id, lab_name, value, test_date, in_range, created_at) "
                "VALUES (?,'INR',?,?,?,?)",
                (patient_id, value, date_string, 1 if 2.0 <= value <= 3.0 else 0, now()),
            )


# --- enrolment --------------------------------------------------------------
def test_enroll_creates_participant(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    with read_db() as conn:
        row = research.get_participant(conn, participant_id)
    assert row["arm"] == "intervention"
    assert row["status"] == "enrolled"
    assert row["baseline_start"] == study_dates["baseline_start"]


def test_enroll_is_idempotent_per_patient(patient, study_dates):
    first = _enroll(patient, study_dates)
    second = _enroll(patient, study_dates, arm="control")
    assert first == second
    with read_db() as conn:
        assert research.get_participant(conn, first)["arm"] == "control"


def test_enroll_requires_consent_date(patient, study_dates):
    with pytest.raises(research.ResearchError, match="ยินยอม"):
        _enroll(patient, study_dates, consent_date="")


def test_screening_status_does_not_require_consent(patient, study_dates):
    participant_id = _enroll(patient, study_dates, status="screening", consent_date="")
    with read_db() as conn:
        assert research.get_participant(conn, participant_id)["status"] == "screening"


def test_enroll_rejects_unknown_arm(patient, study_dates):
    with pytest.raises(research.ResearchError):
        _enroll(patient, study_dates, arm="placebo")


def test_phase_windows_are_derived_when_left_blank():
    windows = research.phase_windows("2026-01-01", {"phase_days": "30"})
    assert windows["baseline_end"] == "2026-01-30"
    assert windows["endline_start"] == "2026-01-31"
    assert windows["endline_end"] == "2026-03-01"


def test_phase_windows_reject_reversed_dates():
    with pytest.raises(research.ResearchError):
        research.phase_windows("2026-02-01", {"baseline_end": "2026-01-01"})


def test_withdrawal_records_reason(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    with db() as conn:
        assert research.withdraw(conn, participant_id, "ย้ายสถานพยาบาล", "pytest")
        row = research.get_participant(conn, participant_id)
    assert row["status"] == "withdrawn"
    assert row["withdrawal_reason"] == "ย้ายสถานพยาบาล"


def test_withdrawal_requires_a_reason(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    with pytest.raises(research.ResearchError), db() as conn:
        research.withdraw(conn, participant_id, "  ", "pytest")


def test_withdrawn_participants_leave_the_analysis_set(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    with read_db() as conn:
        before = len(research.analysis_population(conn))
    with db() as conn:
        research.withdraw(conn, participant_id, "ผู้ป่วยขอถอนตัว", "pytest")
    with read_db() as conn:
        assert len(research.analysis_population(conn)) == before - 1


# --- phase outcomes ---------------------------------------------------------
def test_phase_outcome_computes_adherence_from_dose_rows(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    _seed_doses(patient["patient_id"], study_dates["baseline_start"],
                study_dates["baseline_end"], 0.5)
    _seed_doses(patient["patient_id"], study_dates["endline_start"],
                study_dates["endline_end"], 0.9)
    with read_db() as conn:
        participant = research.get_participant(conn, participant_id)
        baseline = research.phase_outcome(conn, participant, "baseline")
        endline = research.phase_outcome(conn, participant, "endline")
    assert baseline.doses_total == 30
    assert 45 <= baseline.adherence_percent <= 55
    assert 85 <= endline.adherence_percent <= 95
    assert endline.adherence_percent > baseline.adherence_percent


def test_phase_outcome_counts_past_planned_doses_as_missed(patient, study_dates):
    """A historical window has no 'still pending' doses."""
    participant_id = _enroll(patient, study_dates)
    from datetime import timedelta

    from warfarin.time_utils import DATE_FMT, parse_date

    start = parse_date(study_dates["baseline_start"])
    with db() as conn:
        for offset in range(4):
            conn.execute(
                "INSERT INTO medication_plan (patient_id, scheduled_date, scheduled_time, "
                "warfarin_mg, status, created_at) VALUES (?,?,?,?,?,?)",
                (patient["patient_id"], (start + timedelta(days=offset)).strftime(DATE_FMT),
                 "18:00", 3.0, "planned" if offset < 2 else "taken", now()),
            )
    with read_db() as conn:
        participant = research.get_participant(conn, participant_id)
        outcome = research.phase_outcome(conn, participant, "baseline")
    assert outcome.doses_total == 4
    assert outcome.adherence_percent == 50.0


def test_phase_outcome_computes_ttr_and_inr_control(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    _seed_labs(patient["patient_id"], [
        (study_dates["baseline_start"], 1.4),
        (study_dates["baseline_end"], 1.6),
        (study_dates["endline_start"], 2.5),
        (study_dates["endline_end"], 2.7),
    ])
    with read_db() as conn:
        participant = research.get_participant(conn, participant_id)
        baseline = research.phase_outcome(conn, participant, "baseline")
        endline = research.phase_outcome(conn, participant, "endline")
    assert baseline.ttr_percent == 0.0
    assert endline.ttr_percent == 100.0
    assert baseline.inr_in_range_percent == 0.0
    assert endline.inr_in_range_percent == 100.0
    assert endline.mean_inr == 2.6


def test_phase_outcome_without_dates_is_empty(patient):
    with read_db() as conn:
        outcome = research.phase_outcome(
            conn, {"patient_id": patient["patient_id"]}, "baseline"
        )
    assert outcome.doses_total == 0
    assert outcome.adherence_percent is None


# --- instruments ------------------------------------------------------------
def test_record_measurement_and_overwrite(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    with db() as conn:
        research.record_measurement(conn, participant_id, "baseline", "knowledge", 12, "pytest")
        research.record_measurement(conn, participant_id, "baseline", "knowledge", 15, "pytest")
    with read_db() as conn:
        recorded = research.measurements_for(conn, participant_id)
        count = conn.execute(
            "SELECT COUNT(*) FROM study_measurements WHERE participant_id=?",
            (participant_id,),
        ).fetchone()[0]
    assert recorded["knowledge"]["baseline"]["value"] == 15
    assert count == 1, "re-entering a score must replace it, not duplicate it"


def test_measurement_rejects_out_of_range_score(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    with pytest.raises(research.ResearchError), db() as conn:
        research.record_measurement(conn, participant_id, "baseline", "knowledge", 99, "pytest")


def test_measurement_rejects_unknown_instrument(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    with pytest.raises(research.ResearchError), db() as conn:
        research.record_measurement(conn, participant_id, "baseline", "horoscope", 3, "pytest")


def test_measurement_gaps_lists_missing_instruments(patient, study_dates):
    participant_id = _enroll(patient, study_dates)
    with read_db() as conn:
        gaps = research.open_measurement_gaps(conn)
    entry = next(g for g in gaps if g["participant"]["participant_id"] == participant_id)
    assert len(entry["missing"]) == len(research.INSTRUMENTS) * len(research.PHASES)


# --- cohort summaries -------------------------------------------------------
def test_enrollment_summary_counts_by_status_and_arm(patient, study_dates, conn):
    _enroll(patient, study_dates)
    with read_db() as read:
        summary = research.enrollment_summary(read)
    assert summary["by_status"]["enrolled"] >= 1
    assert summary["by_arm"]["intervention"] >= 1
    assert summary["analysed"] >= 1


def test_study_progress_classifies_windows(patient, study_dates):
    _enroll(patient, study_dates)
    with read_db() as conn:
        progress = research.study_progress(conn)
    assert progress["finished"] >= 1


# --- analysis ---------------------------------------------------------------
@pytest.fixture
def cohort(client, study_dates):
    """Six participants with a real improvement between the two phases."""
    import uuid

    from warfarin.patients import clean_patient_form, create_patient

    prefix = uuid.uuid4().hex[:6].upper()
    created = []
    for index in range(6):
        with db() as conn:
            patient_id = create_patient(
                conn,
                clean_patient_form({
                    "full_name": f"ผู้เข้าร่วม {prefix}-{index}",
                    "hn": f"RES-{prefix}-{index:03d}",
                    "age_years": str(60 + index),
                    "sex": "male" if index % 2 else "female",
                    "indication": "af",
                }),
                "pytest",
            )
            patient = dict(
                conn.execute(
                    "SELECT * FROM patients WHERE patient_id=?", (patient_id,)
                ).fetchone()
            )
        arm = "intervention" if index < 4 else "control"
        participant_id = _enroll(patient, study_dates, arm=arm)
        # Vary the individual response: identical improvements everywhere would
        # give a zero SD of differences, which no paired test can evaluate.
        baseline_rate = 0.4 + 0.1 * (index % 3)
        gain = 0.4 - 0.1 * (index % 3) if arm == "intervention" else 0.0
        _seed_doses(patient_id, study_dates["baseline_start"],
                    study_dates["baseline_end"], baseline_rate)
        _seed_doses(patient_id, study_dates["endline_start"],
                    study_dates["endline_end"], baseline_rate + gain)
        with db() as conn:
            research.record_measurement(
                conn, participant_id, "baseline", "knowledge", 10 + index % 3, "pytest"
            )
            research.record_measurement(
                conn, participant_id, "endline", "knowledge", 16 + index % 3, "pytest"
            )
        created.append(participant_id)
    return created


def test_participant_dataset_is_flat_and_complete(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    assert len(dataset) >= 6
    row = dataset[0]
    for key in ("participant_id", "arm", "baseline", "endline", "knowledge_baseline"):
        assert key in row


def test_within_group_analysis_detects_improvement(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    intervention = [row for row in dataset if row["arm"] == "intervention"]
    results = research.within_group_analysis(intervention)
    adherence = next(r for r in results if r["key"] == "adherence_percent")
    assert adherence["endline"].mean > adherence["baseline"].mean
    assert adherence["test"].p_value is not None
    assert adherence["test"].significant


def test_binary_outcome_analysis_uses_mcnemar(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    intervention = [row for row in dataset if row["arm"] == "intervention"]
    results = research.binary_outcome_analysis(intervention)
    adherence = results[0]
    assert adherence["endline"][0] > adherence["baseline"][0]
    assert adherence["test"].name == "McNemar's test"


def test_between_group_analysis_runs_when_both_arms_exist(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    results = research.between_group_analysis(dataset)
    assert results
    adherence = next(r for r in results if r["key"] == "adherence_percent")
    assert adherence["intervention"].mean > adherence["control"].mean


def test_between_group_analysis_is_empty_with_one_arm(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    only_intervention = [row for row in dataset if row["arm"] == "intervention"]
    assert research.between_group_analysis(only_intervention) == []


def test_baseline_characteristics_builds_table_one(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    table = research.baseline_characteristics(dataset)
    assert table["n_total"] >= 6
    assert table["has_control"]
    labels = [row["label"] for row in table["continuous"]]
    assert "อายุ (ปี)" in labels
    categorical = [row["label"] for row in table["categorical"]]
    assert "เพศ" in categorical


# --- export -----------------------------------------------------------------
def test_export_is_deidentified_by_default(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    header, rows = research.export_rows(dataset)
    assert "hn" not in header
    assert "full_name" not in header
    assert "participant_id" in header
    assert len(rows) == len(dataset)


def test_export_can_include_identifiers_when_requested(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    header, _ = research.export_rows(dataset, deidentified=False)
    assert "hn" in header and "full_name" in header


def test_export_includes_change_scores(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    header, rows = research.export_rows(dataset)
    index = header.index("adherence_change")
    assert any(row[index] is not None for row in rows)
    assert "knowledge_change" in header


def test_codebook_matches_export_header(cohort):
    with read_db() as conn:
        dataset = research.participant_dataset(conn)
    header, _ = research.export_rows(dataset)
    names = {entry["name"] for entry in research.codebook()}
    assert set(header) == names


# --- HTTP -------------------------------------------------------------------
def test_research_pages_require_login(anon_client):
    for path in ("/research", "/research/participants", "/research/analysis"):
        assert anon_client.get(path, follow_redirects=False).status_code == 303, path


def test_research_overview_renders(admin_client, cohort):
    response = admin_client.get("/research")
    assert response.status_code == 200
    assert "งานวิจัย" in response.text


def test_participants_page_renders(admin_client, cohort):
    response = admin_client.get("/research/participants")
    assert response.status_code == 200
    assert "ผู้เข้าร่วมวิจัย" in response.text


def test_participants_filter_by_arm(admin_client, cohort):
    response = admin_client.get("/research/participants?arm=control")
    assert response.status_code == 200


def test_participant_detail_renders(admin_client, cohort):
    response = admin_client.get(f"/research/participants/{cohort[0]}")
    assert response.status_code == 200
    assert "ผลลัพธ์ที่คำนวณจากข้อมูลจริง" in response.text


def test_analysis_page_shows_tables(admin_client, cohort):
    response = admin_client.get("/research/analysis")
    assert response.status_code == 200
    assert "ตารางที่ 1" in response.text
    assert "ตารางที่ 2" in response.text
    assert "p-value" in response.text


def test_enroll_form_and_submission_over_http(admin_client, patient, study_dates):
    pid = patient["patient_id"]
    assert admin_client.get(f"/research/enroll/{pid}").status_code == 200
    response = admin_client.post(
        f"/research/enroll/{pid}",
        data={
            "arm": "intervention", "status": "enrolled",
            "consent_date": study_dates["baseline_start"], "consent_version": "v1.0",
            **study_dates,
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    with read_db() as conn:
        assert research.get_by_patient(conn, pid) is not None


def test_enroll_form_reports_validation_error(admin_client, patient, study_dates):
    response = admin_client.post(
        f"/research/enroll/{patient['patient_id']}",
        data={"arm": "intervention", "status": "enrolled", "consent_date": "",
              **study_dates},
    )
    assert response.status_code == 400
    assert "ยินยอม" in response.text


def test_measurement_over_http(admin_client, cohort):
    response = admin_client.post(
        f"/research/participants/{cohort[0]}/measurement",
        data={"instrument": "satisfaction", "phase": "endline", "value": "4"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    with read_db() as conn:
        recorded = research.measurements_for(conn, cohort[0])
    assert recorded["satisfaction"]["endline"]["value"] == 4


def test_withdrawal_over_http(admin_client, cohort):
    admin_client.post(
        f"/research/participants/{cohort[-1]}/status",
        data={"status": "withdrawn", "reason": "เสียชีวิต"},
        follow_redirects=False,
    )
    with read_db() as conn:
        assert research.get_participant(conn, cohort[-1])["status"] == "withdrawn"


def test_csv_export_over_http(admin_client, cohort):
    response = admin_client.get("/research/export.csv")
    assert response.status_code == 200
    text = response.content.decode("utf-8-sig")
    assert "participant_id" in text
    assert "hn" not in text.split("\n")[0].split(",")


def test_identified_export_requires_admin(client, cohort):
    from warfarin.staff import create_staff

    with db() as conn:
        create_staff(conn, "res-nurse", "research-password-1", "พยาบาล", "nurse", "pytest")
    client.cookies.clear()
    client.post("/login", data={"username": "res-nurse", "password": "research-password-1"},
                follow_redirects=False)
    try:
        response = client.get("/research/export.csv?identified=1")
        header = response.content.decode("utf-8-sig").split("\n")[0]
        assert "full_name" not in header, "a nurse must not get identified data"
    finally:
        client.cookies.clear()


def test_admin_can_export_identified(admin_client, cohort):
    response = admin_client.get("/research/export.csv?identified=1")
    header = response.content.decode("utf-8-sig").split("\n")[0]
    assert "full_name" in header


def test_codebook_export_over_http(admin_client):
    response = admin_client.get("/research/codebook.csv")
    assert response.status_code == 200
    assert "variable" in response.content.decode("utf-8-sig")


def test_xlsx_export_has_three_sheets(admin_client, cohort):
    response = admin_client.get("/research/export.xlsx")
    if response.status_code == 501:
        pytest.skip("openpyxl not installed")
    assert response.status_code == 200

    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["data", "codebook", "results"]
    assert workbook["data"].max_row >= 2
    assert workbook["results"].max_row >= 2


def test_nurse_cannot_enroll_participants(client, patient):
    from warfarin.staff import create_staff

    with db() as conn:
        create_staff(conn, "res-nurse2", "research-password-2", "พยาบาล", "nurse", "pytest")
    client.cookies.clear()
    client.post("/login", data={"username": "res-nurse2", "password": "research-password-2"},
                follow_redirects=False)
    try:
        assert client.get(f"/research/enroll/{patient['patient_id']}").status_code == 403
    finally:
        client.cookies.clear()


def test_binary_analysis_reports_zero_denominator(patient, study_dates):
    """With no clinical data at all the denominator is 0, not the numerator."""
    participant_id = _enroll(patient, study_dates)
    with read_db() as conn:
        dataset = [
            row for row in research.participant_dataset(conn)
            if row["participant_id"] == participant_id
        ]
    rows = research.binary_outcome_analysis(dataset)
    for row in rows:
        assert row["baseline_n"] == 0
        assert row["endline_n"] == 0
        assert row["test"].p_value is None


def test_analysis_page_never_shows_a_percentage_with_no_denominator(
    admin_client, patient, study_dates
):
    """A 0% proportion with n=0 must read as 'no data', not 'nobody achieved it'.

    Asserted as an invariant over the rendered table rather than on an empty
    dataset, so the test does not depend on what other tests have inserted.
    """
    import re

    _enroll(patient, study_dates)   # enrolled, but no doses or labs recorded
    response = admin_client.get("/research/analysis")
    assert response.status_code == 200

    section = response.text.split("ตารางที่ 3")[1].split("ตารางที่ 4")[0]
    plain = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", section))
    assert "n=0" not in plain, "a proportion was rendered against an empty denominator"
