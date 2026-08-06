"""Adherence and INR reporting, with CSV / Excel export."""
from __future__ import annotations

import csv
import io
import logging

from fastapi import APIRouter, Depends, Request, Response

from warfarin.adherence import (
    compute_adherence_bulk,
    compute_streak_bulk,
    compute_ttr_bulk,
    last_inr_bulk,
    monthly_trend,
)
from warfarin.audit import log_audit_standalone
from warfarin.clinical import INDICATIONS, assess_inr
from warfarin.db import fetch_all, read_db
from warfarin.deps import csrf_protect, require_user
from warfarin.doses import current_weekly_mg
from warfarin.patients import active_patients
from warfarin.templating import render
from warfarin.time_utils import today

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reports", tags=["reports"], dependencies=[Depends(csrf_protect)])

CSV_HEADERS = [
    "patient_id", "hn", "full_name", "sex", "age_years", "indication",
    "target_inr_min", "target_inr_max", "weekly_mg",
    "doses_7d", "taken_7d", "missed_7d", "adherence_7d_%",
    "doses_30d", "taken_30d", "missed_30d", "adherence_30d_%",
    "streak_days", "last_inr", "last_inr_date", "last_inr_in_range",
    "ttr_%", "next_inr_date",
]


def _build_rows(conn) -> list[dict]:
    """Assemble the per-patient report rows using bulk queries only."""
    patients = active_patients(conn)
    ids = [p["patient_id"] for p in patients]
    adherence7 = compute_adherence_bulk(conn, ids, 7)
    adherence30 = compute_adherence_bulk(conn, ids, 30)
    streaks = compute_streak_bulk(conn, ids)
    ttrs = compute_ttr_bulk(conn, ids)
    last_inr = last_inr_bulk(conn, ids)
    rows = []
    for patient in patients:
        pid = patient["patient_id"]
        latest = last_inr.get(pid)
        rows.append({
            "patient": patient,
            "adh7": adherence7.get(pid),
            "adh30": adherence30.get(pid),
            "streak": streaks.get(pid, 0),
            "ttr": ttrs.get(pid),
            "last_inr": latest,
            "assessment": (
                assess_inr(
                    latest["value"], patient["target_inr_min"], patient["target_inr_max"]
                )
                if latest else None
            ),
            "weekly_mg": current_weekly_mg(conn, pid),
        })
    return rows


@router.get("")
def reports_page(request: Request, user: dict = Depends(require_user)):
    with read_db() as conn:
        rows = _build_rows(conn)
        trend = monthly_trend(conn, months=6)
        survey = fetch_all(
            conn,
            "SELECT AVG(ease_of_use) AS ease, AVG(line_satisfaction) AS line_sat, "
            "AVG(reminder_helpful) AS remind, COUNT(*) AS total FROM satisfaction_surveys",
        )
        scores = fetch_all(
            conn,
            "SELECT test_type, AVG(score * 100.0 / NULLIF(max_score,0)) AS avg_pct, "
            "COUNT(*) AS n FROM test_scores GROUP BY test_type",
        )

    measured = [r for r in rows if r["adh30"] and r["adh30"]["total"]]
    ttr_values = [r["ttr"] for r in rows if r["ttr"] is not None]
    in_range = sum(1 for r in rows if r["last_inr"] and r["last_inr"]["in_range"])
    with_inr = sum(1 for r in rows if r["last_inr"])

    summary = {
        "patients": len(rows),
        "avg_adherence_30d": (
            round(sum(r["adh30"]["percent"] for r in measured) / len(measured), 1)
            if measured else 0
        ),
        "avg_ttr": round(sum(ttr_values) / len(ttr_values), 1) if ttr_values else None,
        "inr_in_range_pct": round(in_range / with_inr * 100, 1) if with_inr else None,
        "patients_with_inr": with_inr,
        "adherence_above_80": sum(
            1 for r in measured if r["adh30"]["percent"] >= 80
        ),
    }
    return render(request, "reports.html", {
        "user": user,
        "report_data": rows,
        "summary": summary,
        "trend": trend,
        "survey_summary": survey[0] if survey else None,
        "test_scores": scores,
        "indications": INDICATIONS,
    })


@router.get("/export.csv")
def export_csv(request: Request, user: dict = Depends(require_user)):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CSV_HEADERS)
    with read_db() as conn:
        rows = _build_rows(conn)
    for row in rows:
        patient = row["patient"]
        latest = row["last_inr"]
        writer.writerow([
            patient["patient_id"], patient.get("hn") or "", patient["full_name"],
            patient.get("sex") or "", patient.get("age_years") or "",
            patient.get("indication") or "",
            patient.get("target_inr_min"), patient.get("target_inr_max"),
            row["weekly_mg"],
            row["adh7"]["total"], row["adh7"]["taken"] + row["adh7"]["late"],
            row["adh7"]["missed"], row["adh7"]["percent"],
            row["adh30"]["total"], row["adh30"]["taken"] + row["adh30"]["late"],
            row["adh30"]["missed"], row["adh30"]["percent"],
            row["streak"],
            latest["value"] if latest else "",
            latest["test_date"] if latest else "",
            ("yes" if latest["in_range"] else "no") if latest else "",
            row["ttr"] if row["ttr"] is not None else "",
            patient.get("next_inr_date") or "",
        ])
    log_audit_standalone("export_report", "reports", "csv", user["username"], "")
    filename = f"warfarin_report_{today()}.csv"
    return Response(
        content=output.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export.xlsx")
def export_xlsx(request: Request, user: dict = Depends(require_user)):
    """Colour-coded Excel export (mirrors the TB tracker's report export)."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed; Excel export unavailable")
        return Response(
            content="ต้องติดตั้ง openpyxl ก่อนจึงจะดาวน์โหลดไฟล์ Excel ได้",
            status_code=501,
            media_type="text/plain; charset=utf-8",
        )

    with read_db() as conn:
        rows = _build_rows(conn)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Adherence"
    headers = [
        "HN", "ชื่อ-นามสกุล", "เพศ", "อายุ", "ข้อบ่งใช้", "เป้าหมาย INR",
        "ขนาดยา/สัปดาห์ (mg)", "โดส 30 วัน", "กินแล้ว", "ขาด",
        "% ความสม่ำเสมอ 30 วัน", "ต่อเนื่อง (วัน)", "INR ล่าสุด",
        "วันที่ตรวจ", "อยู่ในเป้าหมาย", "TTR %", "นัดตรวจครั้งถัดไป",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    green = PatternFill("solid", fgColor="C8E6C9")
    yellow = PatternFill("solid", fgColor="FFF9C4")
    red = PatternFill("solid", fgColor="FFCDD2")

    for row in rows:
        patient = row["patient"]
        latest = row["last_inr"]
        adherence = row["adh30"]
        sheet.append([
            patient.get("hn") or "", patient["full_name"],
            {"male": "ชาย", "female": "หญิง"}.get(patient.get("sex") or "", ""),
            patient.get("age_years") or "",
            INDICATIONS.get(patient.get("indication") or "", {}).get("label", ""),
            f"{patient.get('target_inr_min')}–{patient.get('target_inr_max')}",
            row["weekly_mg"],
            adherence["total"], adherence["taken"] + adherence["late"],
            adherence["missed"], adherence["percent"], row["streak"],
            latest["value"] if latest else "",
            latest["test_date"] if latest else "",
            ("ใช่" if latest["in_range"] else "ไม่") if latest else "",
            row["ttr"] if row["ttr"] is not None else "",
            patient.get("next_inr_date") or "",
        ])
        current_row = sheet.max_row
        if adherence["total"]:
            percent_cell = sheet.cell(row=current_row, column=11)
            percent_cell.fill = (
                green if adherence["percent"] >= 90
                else yellow if adherence["percent"] >= 70
                else red
            )
        if latest:
            inr_cell = sheet.cell(row=current_row, column=13)
            inr_cell.fill = green if latest["in_range"] else red

    widths = [12, 26, 8, 8, 26, 14, 16, 12, 10, 8, 18, 14, 12, 14, 14, 10, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(index)
        ].width = width
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    log_audit_standalone("export_report", "reports", "xlsx", user["username"], "")
    filename = f"warfarin_report_{today()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
