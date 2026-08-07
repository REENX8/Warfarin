"""Research study console: cohort management, analysis and data export."""
from __future__ import annotations

import csv
import io
import logging

from fastapi import APIRouter, Depends, HTTPException, Request, Response

from warfarin import research
from warfarin.audit import log_audit_standalone
from warfarin.db import db, read_db
from warfarin.deps import csrf_protect, redirect, require_role, require_user
from warfarin.patients import get_patient
from warfarin.templating import render
from warfarin.time_utils import today

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/research", tags=["research"], dependencies=[Depends(csrf_protect)])

# Enrolling and un-enrolling participants is a study-conduct decision.
manage_study = require_role("admin", "pharmacist")


@router.get("")
def overview(request: Request, study: str = "", user: dict = Depends(require_user)):
    with read_db() as conn:
        summary = research.enrollment_summary(conn, study)
        progress = research.study_progress(conn, study)
        gaps = research.open_measurement_gaps(conn, study)
        codes = research.study_codes(conn)
        participants = research.list_participants(conn, study_code=study)
    return render(request, "research_overview.html", {
        "user": user,
        "summary": summary,
        "progress": progress,
        "gaps": gaps[:20],
        "gap_count": len(gaps),
        "study": study,
        "study_codes": codes,
        "participants": participants[:12],
        "arms": research.ARMS,
        "statuses": research.STATUSES,
    })


@router.get("/participants")
def participants_page(
    request: Request,
    arm: str = "",
    status: str = "",
    study: str = "",
    user: dict = Depends(require_user),
):
    with read_db() as conn:
        rows = research.list_participants(conn, arm=arm, status=status, study_code=study)
        measurements = research.measurements_bulk(
            conn, [row["participant_id"] for row in rows]
        )
        codes = research.study_codes(conn)
    for row in rows:
        recorded = measurements.get(row["participant_id"], {})
        row["scores_recorded"] = sum(
            1 for instrument in research.INSTRUMENTS
            for phase in research.PHASES
            if recorded.get(instrument, {}).get(phase) is not None
        )
    total_slots = len(research.INSTRUMENTS) * len(research.PHASES)
    return render(request, "research_participants.html", {
        "user": user,
        "participants": rows,
        "arm": arm,
        "status": status,
        "study": study,
        "study_codes": codes,
        "arms": research.ARMS,
        "statuses": research.STATUSES,
        "total_slots": total_slots,
    })


@router.get("/enroll/{patient_id}")
def enroll_form(
    request: Request, patient_id: int, user: dict = Depends(manage_study)
):
    with read_db() as conn:
        patient = get_patient(conn, patient_id)
        if patient is None:
            raise HTTPException(404, "ไม่พบผู้ป่วย")
        participant = research.get_by_patient(conn, patient_id)
        codes = research.study_codes(conn)
    return render(request, "research_enroll.html", {
        "user": user,
        "patient": patient,
        "participant": participant,
        "arms": research.ARMS,
        "statuses": research.STATUSES,
        "study_codes": codes,
        "default_code": research.DEFAULT_STUDY_CODE,
        "default_phase_days": research.DEFAULT_PHASE_DAYS,
        "today": today(),
        "error": "",
    })


@router.post("/enroll/{patient_id}")
async def enroll_submit(
    request: Request, patient_id: int, user: dict = Depends(manage_study)
):
    form = dict(await request.form())
    with read_db() as conn:
        patient = get_patient(conn, patient_id)
        participant = research.get_by_patient(conn, patient_id)
        codes = research.study_codes(conn)
    if patient is None:
        raise HTTPException(404, "ไม่พบผู้ป่วย")
    try:
        with db() as conn:
            participant_id = research.enroll(conn, patient_id, form, user["username"])
    except research.ResearchError as exc:
        return render(request, "research_enroll.html", {
            "user": user, "patient": patient, "participant": participant,
            "arms": research.ARMS, "statuses": research.STATUSES,
            "study_codes": codes, "default_code": research.DEFAULT_STUDY_CODE,
            "default_phase_days": research.DEFAULT_PHASE_DAYS,
            "today": today(), "error": str(exc), "form": form,
        }, status_code=400)
    return redirect(
        f"/research/participants/{participant_id}", "บันทึกข้อมูลการเข้าร่วมวิจัยเรียบร้อย"
    )


@router.get("/participants/{participant_id}")
def participant_detail(
    request: Request, participant_id: int, user: dict = Depends(require_user)
):
    with read_db() as conn:
        participant = research.get_participant(conn, participant_id)
        if participant is None:
            raise HTTPException(404, "ไม่พบผู้เข้าร่วมวิจัย")
        baseline = research.phase_outcome(conn, participant, "baseline")
        endline = research.phase_outcome(conn, participant, "endline")
        measurements = research.measurements_for(conn, participant_id)
    return render(request, "research_participant.html", {
        "user": user,
        "participant": participant,
        "baseline": baseline,
        "endline": endline,
        "measurements": measurements,
        "instruments": research.INSTRUMENTS,
        "phases": research.PHASES,
        "arms": research.ARMS,
        "statuses": research.STATUSES,
        "withdrawal_reasons": research.WITHDRAWAL_REASONS,
        "today": today(),
    })


@router.post("/participants/{participant_id}/measurement")
async def record_measurement(
    request: Request, participant_id: int, user: dict = Depends(manage_study)
):
    form = await request.form()
    try:
        with db() as conn:
            research.record_measurement(
                conn,
                participant_id=participant_id,
                phase=str(form.get("phase") or ""),
                instrument=str(form.get("instrument") or ""),
                value=form.get("value"),
                performed_by=user["username"],
                measured_on=str(form.get("measured_on") or ""),
                text_value=str(form.get("text_value") or ""),
            )
    except research.ResearchError as exc:
        return redirect(f"/research/participants/{participant_id}", str(exc), "danger")
    return redirect(f"/research/participants/{participant_id}", "บันทึกคะแนนเรียบร้อย")


@router.post("/participants/{participant_id}/status")
async def update_status(
    request: Request, participant_id: int, user: dict = Depends(manage_study)
):
    form = await request.form()
    status = str(form.get("status") or "")
    try:
        with db() as conn:
            if status == "withdrawn":
                ok = research.withdraw(
                    conn, participant_id, str(form.get("reason") or ""), user["username"]
                )
            else:
                ok = research.set_status(conn, participant_id, status, user["username"])
    except research.ResearchError as exc:
        return redirect(f"/research/participants/{participant_id}", str(exc), "danger")
    if not ok:
        return redirect("/research/participants", "ไม่พบผู้เข้าร่วมวิจัย", "danger")
    return redirect(f"/research/participants/{participant_id}", "อัปเดตสถานะเรียบร้อย")


@router.get("/analysis")
def analysis(request: Request, study: str = "", user: dict = Depends(require_user)):
    with read_db() as conn:
        dataset = research.participant_dataset(conn, study)
        summary = research.enrollment_summary(conn, study)
        codes = research.study_codes(conn)
    intervention = [row for row in dataset if row["arm"] == "intervention"]
    return render(request, "research_analysis.html", {
        "user": user,
        "study": study,
        "study_codes": codes,
        "summary": summary,
        "n_analysed": len(dataset),
        "table_one": research.baseline_characteristics(dataset),
        "within": research.within_group_analysis(intervention or dataset),
        "within_label": (
            "กลุ่มทดลอง" if intervention and len(intervention) != len(dataset)
            else "ผู้เข้าร่วมทั้งหมด"
        ),
        "binary": research.binary_outcome_analysis(intervention or dataset),
        "between": research.between_group_analysis(dataset),
        "instruments": research.INSTRUMENTS,
        "adherence_target": research.ADHERENCE_TARGET,
        "ttr_target": research.TTR_TARGET,
    })


# ---------------------------------------------------------------------------
# Data export
# ---------------------------------------------------------------------------
def _dataset(study: str) -> list[dict]:
    with read_db() as conn:
        return research.participant_dataset(conn, study)


@router.get("/export.csv")
def export_csv(
    request: Request,
    study: str = "",
    identified: int = 0,
    user: dict = Depends(require_user),
):
    """Analysis dataset as CSV. De-identified unless explicitly requested."""
    deidentified = not (identified and user["role"] == "admin")
    header, rows = research.export_rows(_dataset(study), deidentified=deidentified)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(header)
    for row in rows:
        writer.writerow(["" if value is None else value for value in row])

    log_audit_standalone(
        "export_research", "research", study or "all", user["username"],
        f"format=csv deidentified={deidentified}",
    )
    suffix = "deidentified" if deidentified else "identified"
    filename = f"warfarin_study_{suffix}_{today()}.csv"
    return Response(
        content=output.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/codebook.csv")
def export_codebook(
    request: Request, identified: int = 0, user: dict = Depends(require_user)
):
    """Variable dictionary that goes with the dataset."""
    deidentified = not (identified and user["role"] == "admin")
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["variable", "label", "type"])
    for entry in research.codebook(deidentified=deidentified):
        writer.writerow([entry["name"], entry["label"], entry["type"]])
    return Response(
        content=output.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="warfarin_codebook_{today()}.csv"'
        },
    )


@router.get("/export.xlsx")
def export_xlsx(
    request: Request,
    study: str = "",
    identified: int = 0,
    user: dict = Depends(require_user),
):
    """Workbook with the data, the codebook and the analysis tables."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(
            content="ต้องติดตั้ง openpyxl ก่อนจึงจะดาวน์โหลดไฟล์ Excel ได้",
            status_code=501, media_type="text/plain; charset=utf-8",
        )

    deidentified = not (identified and user["role"] == "admin")
    dataset = _dataset(study)
    header, rows = research.export_rows(dataset, deidentified=deidentified)

    workbook = openpyxl.Workbook()
    data_sheet = workbook.active
    data_sheet.title = "data"
    data_sheet.append(header)
    for row in rows:
        data_sheet.append(row)
    for cell in data_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    data_sheet.freeze_panes = "A2"
    for index, name in enumerate(header, start=1):
        data_sheet.column_dimensions[get_column_letter(index)].width = max(14, len(name) + 2)

    codebook_sheet = workbook.create_sheet("codebook")
    codebook_sheet.append(["variable", "label", "type"])
    for entry in research.codebook(deidentified=deidentified):
        codebook_sheet.append([entry["name"], entry["label"], entry["type"]])
    for cell in codebook_sheet[1]:
        cell.font = Font(bold=True)
    codebook_sheet.column_dimensions["A"].width = 30
    codebook_sheet.column_dimensions["B"].width = 56
    codebook_sheet.column_dimensions["C"].width = 12

    intervention = [row for row in dataset if row["arm"] == "intervention"] or dataset
    results_sheet = workbook.create_sheet("results")
    results_sheet.append([
        "ตัวชี้วัด", "n", "Baseline (mean ± SD)", "Endline (mean ± SD)",
        "ผลต่างเฉลี่ย", "t", "df", "p-value", "Effect size",
    ])
    for cell in results_sheet[1]:
        cell.font = Font(bold=True)
    for entry in research.within_group_analysis(intervention):
        test = entry["test"]
        results_sheet.append([
            entry["label"], test.n,
            entry["baseline"].summary(), entry["endline"].summary(),
            round(test.extra.get("mean_difference"), 2)
            if test.extra.get("mean_difference") is not None else None,
            round(test.statistic, 3) if test.statistic is not None else None,
            test.df,
            test.p_text(),
            round(test.effect_size, 3) if test.effect_size is not None else None,
        ])
    results_sheet.column_dimensions["A"].width = 42
    for column in "CDE":
        results_sheet.column_dimensions[column].width = 22

    buffer = io.BytesIO()
    workbook.save(buffer)
    log_audit_standalone(
        "export_research", "research", study or "all", user["username"],
        f"format=xlsx deidentified={deidentified}",
    )
    suffix = "deidentified" if deidentified else "identified"
    filename = f"warfarin_study_{suffix}_{today()}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
